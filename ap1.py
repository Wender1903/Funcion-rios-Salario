import os
import pyautogui
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from porcentagemDeAumento import porcentagem_de_aumento
from calcularAumento import calcular_aumento
        
        
funcionarios = {
    'Nome': ['Wender', 'Jose', 'Tania', 'Rillary'],
    'Cargo': ['Tecnico', 'Especialista', 'Operador', 'Assistente'],
    'Salario': [3700.0, 7500.0, 2800.0, 1700.0],
    'Anos de Serviço': [3, 7, 8, 2]
}

df = pd.DataFrame(funcionarios)
print(df)

df['Aumento'] = df.apply(porcentagem_de_aumento, axis=1)
df['Salario Aumentado'] = df.apply(calcular_aumento, axis=1)
print(df)

arquivo_excel = 'funcionarios.xlsx'
df.to_excel(arquivo_excel, index=False, sheet_name='Funcionarios')

wb = load_workbook(arquivo_excel)
ws = wb['Funcionarios']

header_fill = PatternFill(start_color="4e97d1", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True, color="ffffff")
    cell.fill = header_fill
    

for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
    for cell in row:
        if cell.column in [3, 6]:
            cell.number_format = '0.00'
            
            
for row in ws.iter_rows(min_row=2, min_col=2, max_col=6):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            if cell.value > 1500 and cell.value < 2000:
                cell.font = Font(color="ff0000")
            elif cell.value > 2500 and cell.value < 3500:
                cell.font = Font(color="ffa07a")
            elif cell.value > 3500 and cell.value < 5500:
                cell.font = Font(color="ffcc66")
            elif cell.value > 5500 and cell.value < 10000:
                cell.font = Font(color="90ee90")
                
wb.save(arquivo_excel)
print(f'A planilha "{arquivo_excel}" foi criada e formatada com sucesso!!')
            
